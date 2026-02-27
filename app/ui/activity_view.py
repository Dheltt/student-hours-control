from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QDateEdit,
    QHeaderView, QMessageBox, QStackedWidget,
    QFileDialog, QFrame
)
from PySide6.QtCore import Qt, QDate
from PySide6 import QtCore
from PySide6.QtGui import QIntValidator, QRegularExpressionValidator
from PySide6.QtCore import QRegularExpression
QtCore.qInstallMessageHandler(lambda msg_type, context, message: None)

from datetime import datetime, timedelta

from app.services.activity_service import ActivityService


class ActividadView(QWidget):

    def __init__(self):
        super().__init__()

        # 🔥 Estado interno para fila seleccionada
        self.selected_activity_id = None

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(30, 30, 30, 30)
        main_layout.setSpacing(20)

        # =====================================================
        # BOTONES DE NAVEGACIÓN
        # =====================================================
        nav_layout = QHBoxLayout()

        self.btn_crud = QPushButton("Gestionar Actividades")
        self.btn_registro = QPushButton("Consultar Registros")

        self.btn_crud.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        self.btn_registro.clicked.connect(lambda: self.stack.setCurrentIndex(1))

        nav_layout.addWidget(self.btn_crud)
        nav_layout.addWidget(self.btn_registro)
        nav_layout.addStretch()

        main_layout.addLayout(nav_layout)

        # =====================================================
        # STACK
        # =====================================================
        self.stack = QStackedWidget()

        # =====================================================
        # VISTA 1 — CRUD ACTIVIDADES
        # =====================================================
        self.view_crud = QWidget()
        crud_layout = QVBoxLayout(self.view_crud)

        form_layout = QHBoxLayout()

        self.nombre_input = QLineEdit()
        self.nombre_input.setPlaceholderText("Nombre actividad")
        regex_nombre = QRegularExpression("^[A-Za-zÁÉÍÓÚáéíóúÑñ ]+$")
        validator_nombre = QRegularExpressionValidator(regex_nombre)
        self.nombre_input.setValidator(validator_nombre)

        self.btn_add = QPushButton("Agregar")
        self.btn_update = QPushButton("Actualizar")
        self.btn_delete = QPushButton("Eliminar")

        self.btn_add.clicked.connect(self.add_activity)
        self.btn_update.clicked.connect(self.update_activity)
        self.btn_delete.clicked.connect(self.delete_activity)

        form_layout.addWidget(self.nombre_input)
        form_layout.addWidget(self.btn_add)
        form_layout.addWidget(self.btn_update)
        form_layout.addWidget(self.btn_delete)

        crud_layout.addLayout(form_layout)

        self.table_crud = self.create_table(["ID", "Nombre Actividad"])
        crud_layout.addWidget(self.table_crud)

        #Conexión profesional
        self.table_crud.itemSelectionChanged.connect(self.load_selected_activity)

        # =====================================================
        # VISTA 2 — CONSULTA REGISTROS
        # =====================================================
        self.view_registro = QWidget()
        registro_layout = QVBoxLayout(self.view_registro)

        filter_layout = QHBoxLayout()
        
        self.lbl_from = QLabel("Fecha Inicio:")
        self.lbl_to = QLabel("Fecha Fin:")

        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_from.setStyleSheet("background-color: white; color: black;")

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setStyleSheet("background-color: white; color: black;")

        self.btn_filter = QPushButton("Filtrar")
        self.btn_today = QPushButton("Hoy")
        self.btn_week = QPushButton("Esta Semana")
        self.btn_last_week = QPushButton("Semana Pasada")
        self.btn_month = QPushButton("Este Mes")

        self.btn_filter.clicked.connect(self.filter_by_range)
        self.btn_today.clicked.connect(self.filter_today)
        self.btn_week.clicked.connect(self.filter_week)
        self.btn_last_week.clicked.connect(self.filter_last_week)
        self.btn_month.clicked.connect(self.filter_month)
        

        filter_layout.addWidget(self.lbl_from)
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(self.lbl_to)
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(self.btn_filter)
        filter_layout.addWidget(self.btn_today)
        filter_layout.addWidget(self.btn_week)
        filter_layout.addWidget(self.btn_last_week)
        filter_layout.addWidget(self.btn_month)

        registro_layout.addLayout(filter_layout)
        
        # =====================================================
        # TABLAS CON ENCABEZADO PROFESIONAL
        # =====================================================
        self.resumen_frame, self.table_resumen = self.create_table_section("Resumen de Actividades")
        self.detalle_frame, self.table_detalle = self.create_table_section("Detalle de Alumnos")

        registro_layout.addWidget(self.resumen_frame)
        registro_layout.addWidget(self.detalle_frame)

        # Botones exportación
        export_layout = QHBoxLayout()

        self.btn_export_pdf = QPushButton("Exportar PDF")
        self.btn_export_excel = QPushButton("Exportar Excel")

        self.btn_export_pdf.clicked.connect(self.export_pdf)
        self.btn_export_excel.clicked.connect(self.export_excel)

        # =================================
        # ESTILOS DE BOTONES (después de crearlos)
        # =================================

        # CRUD Actividades
        self.btn_add.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color:white;
                border-radius: 10px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)

        self.btn_update.setStyleSheet("""
            QPushButton {
                background-color: #c7d2fe;
                color:#1e3a8a;
                border-radius: 10px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #a5b4fc;
            }
        """)

        self.btn_delete.setStyleSheet("""
            QPushButton {
                background-color: #f87171;
                color:white;
                border-radius: 10px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #ef4444;
            }
        """)

        # Navegación y filtros
        for btn in [self.btn_crud, self.btn_registro, self.btn_filter, self.btn_today, self.btn_week, self.btn_last_week, self.btn_month]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #4f46e5;
                    color:white;
                    border-radius: 8px;
                    padding: 6px 12px;
                }
                QPushButton:hover {
                    background-color: #4338ca;
                }
            """)

        # Exportación
        self.btn_export_pdf.setStyleSheet("""
            QPushButton {
                background-color: #f87171;
                color: white;
                border-radius: 10px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #ef4444;
            }
        """)
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #22c55e;
                color:white;
                border-radius: 10px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #16a34a;
            }
        """)

        export_layout.addStretch()
        export_layout.addWidget(self.btn_export_pdf)
        export_layout.addWidget(self.btn_export_excel)

        registro_layout.addLayout(export_layout)

        # =====================================================
        # AGREGAR VISTAS AL STACK
        # =====================================================
        self.stack.addWidget(self.view_crud)
        self.stack.addWidget(self.view_registro)

        main_layout.addWidget(self.stack)

        # Cargar datos iniciales
        self.refresh_crud()
        self.filter_by_range()

    # =====================================================
    # MENSAJE CLARO
    # =====================================================
    def show_light_message(self, title, message, icon):
        msg = QMessageBox(self)
        msg.setIcon(icon)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStyleSheet("""
            QMessageBox { background-color: white; }
            QLabel { color: gray; }
            QPushButton { background-color: #f0f0f0; padding: 5px 15px; }
        """)
        msg.exec()

    # =====================================================
    # TABLA (TUS ESTILOS)
    # =====================================================
    def create_table(self, headers):
        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        table.setStyleSheet("""
            QTableWidget {
                background-color: #f4f6f9;
                border: 1px solid #cfd6df;
                border-radius: 18px;
                gridline-color: transparent;
            }
            QHeaderView::section {
                background-color: #dde3ea;
                padding: 10px;
                border: none;
                font-weight: 600;
                color: #2f3437;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e3e8ef;
            }
            QTableWidget::item:selected {
                background-color: #dbeafe;
                color: #1e3a8a;
            }
        """)

        return table

    def create_table_section(self, title):
        """Crea un contenedor para la tabla con encabezado profesional."""
        frame = QFrame()
        layout = QVBoxLayout(frame)
        
        lbl_title = QLabel(title)
        lbl_title.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            background-color: #b0b3b8;
            color: 4f46e5;
            padding: 5px 10px;
            border-radius: 5px;
        """)
        
        table = self.create_table([])
        layout.addWidget(lbl_title)
        layout.addWidget(table)
        return frame, table

    # =====================================================
    # CRUD ACTIVIDADES
    # =====================================================
    def refresh_crud(self):
        data = ActivityService.get_all_activities() or []
        self.table_crud.setRowCount(len(data))

        for row, item in enumerate(data):
            for col, value in enumerate(item):
                self.table_crud.setItem(row, col, QTableWidgetItem(str(value)))

        # 🔥 RESET PROFESIONAL
        self.selected_activity_id = None
        self.nombre_input.clear()

    def load_selected_activity(self):
        row = self.table_crud.currentRow()

        if row < 0:
            self.selected_activity_id = None
            return

        id_item = self.table_crud.item(row, 0)
        name_item = self.table_crud.item(row, 1)

        if id_item and name_item:
            self.selected_activity_id = int(id_item.text())
            self.nombre_input.setText(name_item.text())

    def add_activity(self):
        try:
            nombre = self.nombre_input.text().strip()

            if not nombre:
                self.show_light_message("Validación", "El nombre no puede estar vacío.", QMessageBox.Warning)
                return

            ActivityService.add_activity(nombre)
            self.refresh_crud()
            self.show_light_message("Éxito", "Actividad agregada correctamente.", QMessageBox.Information)
        except Exception as e:
            self.show_light_message("Error", str(e), QMessageBox.Critical)

    def update_activity(self):
        if not self.selected_activity_id:
            self.show_light_message("Atención", "Selecciona una actividad.", QMessageBox.Warning)
            return

        nombre = self.nombre_input.text().strip()

        if not nombre:
            self.show_light_message("Validación", "El nombre no puede estar vacío.", QMessageBox.Warning)
            return

        confirm = QMessageBox.question(
            self,
            "Confirmar actualización",
            "¿Estás seguro de modificar esta actividad?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.No:
            return

        try:
            ActivityService.update_activity(self.selected_activity_id, nombre)
            self.refresh_crud()
            self.show_light_message("Éxito", "Actividad actualizada correctamente.", QMessageBox.Information)
        except Exception as e:
            self.show_light_message("Error", str(e), QMessageBox.Critical)

    def delete_activity(self):
        if not self.selected_activity_id:
            self.show_light_message("Atención", "Selecciona una actividad.", QMessageBox.Warning)
            return

        confirm = QMessageBox.question(
            self,
            "Confirmar eliminación",
            "¿Estás seguro de eliminar esta actividad?\nEsta acción no se puede deshacer.",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.No:
            return

        try:
            ActivityService.delete_activity(self.selected_activity_id)
            self.refresh_crud()
            self.show_light_message("Éxito", "Actividad eliminada correctamente.", QMessageBox.Information)
        except Exception as e:
            self.show_light_message("Error", str(e), QMessageBox.Critical)

    # =====================================================
    # FILTROS
    # =====================================================
    def load_registro_data(self):
        """Carga los datos de resumen y detalle en las tablas."""
        inicio = self.date_from.date().toPython()
        fin = self.date_to.date().toPython()

        resumen = ActivityService.get_registros_summary(inicio, fin) or []
        detalle = ActivityService.get_alumnos_por_registro(inicio, fin) or []
        
        # ===== RESUMEN =====
        self.table_resumen.setColumnCount(4)
        self.table_resumen.setHorizontalHeaderLabels(
            ["Actividad", "Fecha", "Horas Totales", "Cantidad Alumnos"]
        )
        self.table_resumen.setRowCount(len(resumen))

        for row, item in enumerate(resumen):
            # item = [actividad, fecha, horas_totales, cantidad_alumnos]
            self.table_resumen.setItem(row, 0, QTableWidgetItem(str(item[0])))
            self.table_resumen.setItem(row, 1, QTableWidgetItem(str(item[1])))
            self.table_resumen.setItem(row, 2, QTableWidgetItem(str(item[2])))
            self.table_resumen.setItem(row, 3, QTableWidgetItem(str(item[3])))

        # ===== DETALLE =====
        self.table_detalle.setColumnCount(4)
        self.table_detalle.setHorizontalHeaderLabels(
            ["Alumno", "Número Control", "Actividad","Fecha"]
        )
        self.table_detalle.setRowCount(len(detalle))

        for row, item in enumerate(detalle):
            # item = (actividad, fecha, numero_control, nombre)
            datos = [item[0], item[1], item[2], item[3]]

            for col, value in enumerate(datos):
                item_widget = QTableWidgetItem(str(value))
                item_widget.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.table_detalle.setItem(row, col, item_widget)

    def filter_by_range(self):
        self.load_registro_data()

    def filter_today(self):
        today = datetime.today().date()
        self.date_from.setDate(QDate(today.year, today.month, today.day))
        self.date_to.setDate(QDate(today.year, today.month, today.day))
        self.load_registro_data()

    def filter_week(self):
        today = datetime.today().date()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        self.date_from.setDate(QDate(start.year, start.month, start.day))
        self.date_to.setDate(QDate(end.year, end.month, end.day))
        self.load_registro_data()

    def filter_last_week(self):
        today = datetime.today().date()
        start_this_week = today - timedelta(days=today.weekday())
        start_last_week = start_this_week - timedelta(days=7)
        end_last_week = start_last_week + timedelta(days=6)
        self.date_from.setDate(QDate(start_last_week.year, start_last_week.month, start_last_week.day))
        self.date_to.setDate(QDate(end_last_week.year, end_last_week.month, end_last_week.day))
        self.load_registro_data()

    def filter_month(self):
        today = datetime.today().date()
        start = today.replace(day=1)
        if today.month == 12:
            end = today.replace(year=today.year+1, month=1, day=1) - timedelta(days=1)
        else:
            end = today.replace(month=today.month+1, day=1) - timedelta(days=1)
        self.date_from.setDate(QDate(start.year, start.month, start.day))
        self.date_to.setDate(QDate(end.year, end.month, end.day))
        self.load_registro_data()

    # =====================================================
    # EXPORTAR
    # =====================================================
    def export_pdf(self):
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet

        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", "", "PDF Files (*.pdf)")
        if not file_path:
            return

        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph("Reporte de Actividades", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Fecha de Consulta: {self.date_from.date().toString('dd/MM/yyyy')} - {self.date_to.date().toString('dd/MM/yyyy')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # ===== RESUMEN =====
        elements.append(Paragraph("Resumen de Actividades", styles['Heading2']))
        resumen_data = [[self.table_resumen.horizontalHeaderItem(i).text() for i in range(self.table_resumen.columnCount())]]
        for row in range(self.table_resumen.rowCount()):
            resumen_data.append([self.table_resumen.item(row, col).text() if self.table_resumen.item(row, col) else "" 
                                 for col in range(self.table_resumen.columnCount())])
        resumen_table = Table(resumen_data, hAlign='LEFT')
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3a8a")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey])
        ]))
        elements.append(resumen_table)
        elements.append(Spacer(1, 20))

        # ===== DETALLE =====
        elements.append(Paragraph("Detalle de Alumnos", styles['Heading2']))
        detalle_data = [[self.table_detalle.horizontalHeaderItem(i).text() for i in range(self.table_detalle.columnCount())]]
        for row in range(self.table_detalle.rowCount()):
            detalle_data.append([self.table_detalle.item(row, col).text() if self.table_detalle.item(row, col) else "" 
                                 for col in range(self.table_detalle.columnCount())])
        detalle_table = Table(detalle_data, hAlign='LEFT')
        detalle_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3a8a")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey])
        ]))
        elements.append(detalle_table)

        doc.build(elements)
        self.show_light_message("Exportación exitosa", f"PDF guardado en:\n{file_path}", QMessageBox.Information)


    def export_excel(self):
        self.load_registro_data()

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Reporte Excel", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")

        inicio = self.date_from.date().toString("dd/MM/yyyy")
        fin = self.date_to.date().toString("dd/MM/yyyy")

        # =====================================================
        # HOJA RESUMEN
        # =====================================================
        ws1 = wb.active
        ws1.title = "Resumen"

        ws1.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=self.table_resumen.columnCount())
        ws1.cell(row=1, column=1, value="REPORTE DE ACTIVIDADES")
        ws1.cell(row=1, column=1).font = Font(size=14, bold=True)

        ws1.cell(row=2, column=1,
                value=f"Periodo: {inicio} - {fin}")
        ws1.cell(row=2, column=1).font = Font(italic=True)

        start_row = 4

        # Encabezados
        for col in range(self.table_resumen.columnCount()):
            cell = ws1.cell(row=start_row, column=col+1,
                            value=self.table_resumen.horizontalHeaderItem(col).text())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Datos
        for row in range(self.table_resumen.rowCount()):
            for col in range(self.table_resumen.columnCount()):
                item = self.table_resumen.item(row, col)
                cell = ws1.cell(row=row+start_row+1, column=col+1,
                                value=item.text() if item else "")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

        ws1.freeze_panes = "A5"
        ws1.auto_filter.ref = f"A4:{get_column_letter(self.table_resumen.columnCount())}{self.table_resumen.rowCount()+4}"

        # Ajustar ancho
        for col in ws1.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # =====================================================
        # HOJA DETALLE
        # =====================================================
        ws2 = wb.create_sheet("Detalle")

        ws2.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=self.table_detalle.columnCount())
        ws2.cell(row=1, column=1, value="DETALLE DE ALUMNOS")
        ws2.cell(row=1, column=1).font = Font(size=14, bold=True)

        ws2.cell(row=2, column=1,
                value=f"Periodo: {inicio} - {fin}")
        ws2.cell(row=2, column=1).font = Font(italic=True)

        start_row = 4

        # Encabezados
        for col in range(self.table_detalle.columnCount()):
            cell = ws2.cell(row=start_row, column=col+1,
                            value=self.table_detalle.horizontalHeaderItem(col).text())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Datos
        for row in range(self.table_detalle.rowCount()):
            for col in range(self.table_detalle.columnCount()):
                item = self.table_detalle.item(row, col)
                cell = ws2.cell(row=row+start_row+1, column=col+1,
                                value=item.text() if item else "")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left")

        ws2.freeze_panes = "A5"
        ws2.auto_filter.ref = f"A4:{get_column_letter(self.table_detalle.columnCount())}{self.table_detalle.rowCount()+4}"

        for col in ws2.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(file_path)

        # =====================================================
        # MENSAJE DE CONFIRMACIÓN
        # =====================================================
        QMessageBox.information(
            self,
            "Exportación exitosa",
            f"El archivo se guardó correctamente en:\n{file_path}"
        ) 